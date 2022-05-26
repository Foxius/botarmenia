from bs4 import BeautifulSoup as bs
import requests 
from pyowm import OWM
from pyowm.utils import config
from pyowm.utils import timestamps
from pyowm.utils.config import get_default_config
import sqlite3
from openpyxl import load_workbook
import openpyxl

config_dict = get_default_config()
owm = OWM('d99de04f9c4d8f439d5149e9d7c689ff', config_dict)
place = 'Ереван'
mgr = owm.weather_manager()
observation = mgr.weather_at_place(place)
w = observation.weather
t = w.temperature("celsius")
t1 = t['temp']
t2 = t['feels_like']
t3 = t['temp_max']
t4 = t['temp_min']

wi = w.wind()['speed']
humi = w.humidity
cl = w.clouds
st = w.status
dt = w.detailed_status
ti = w.reference_time('iso')
pr = w.pressure['press']
vd = w.visibility_distance

press = int(pr) * float("0.75")

textweather = f"""В городе ***Ереван*** сейчас:
	*Температура*: `{str(t1)}°C`
	*Ощущается как*: `{str(t2)}°C`
	*Скорость ветра*: `{str(wi)}м/с`
	*Давление*: `{str(press)} мм.рт.ст`
	*Влажность*: `{str(humi)}%`
	"""		
print(textweather)







# wb = openpyxl.reader.excel.load_workbook(filename="xlsx/basadate.xlsx", data_only=True)
# print(wb.sheetnames)
# wb.active = 1
# sheet = wb.active
# #print(sheet['A1'].value)
# for i in range(1,12):
#     print(sheet['A'+str(i)].value,sheet['B'+str(i)].value,sheet['C'+str(i)].value)







# wb = openpyxl.reader.excel.load_workbook(filename="sample.xlsx", data_only=True)
# print(wb.sheetnames)
# wb.active = 1
# sheet = wb.active
# #print(sheet['A1'].value)
# # for i in range(1,12):
# #     print(sheet['A'+str(i)].value,sheet['B'+str(i)].value,sheet['C'+str(i)].value)




# with sqlite3.connect('maindb.db') as db:
# 	cursor = db.cursor()
# 	query1 = ''' INSERT INTO Shows VALUES('Stranger Things', 'Shawn Levy', 2016) '''
# 	cursor.execute(query1)
# 	db.commit()


# cursor.execute('''CREATE TABLE IF NOT EXISTS Shows
#               (Title TEXT, Director TEXT, Year INT)''')
# config_dict = get_default_config()
# VALUES ('Stranger Things', 'Shawn Levy', 2016)
# owm = OWM('d99de04f9c4d8f439d5149e9d7c689ff', config_dict)
# # config_dict['language'] = 'ru'
# place = 'Дубай'
# mgr = owm.weather_manager()
# observation = mgr.weather_at_place(place)
# w = observation.weather
# t = w.temperature("celsius")
# t1 = t['temp']
# t2 = t['feels_like']
# t3 = t['temp_max']
# t4 = t['temp_min']
# # t5 = t['weather.description']
# wi = w.wind()['speed']
# humi = w.humidity
# cl = w.clouds
# st = w.status
# dt = w.detailed_status
# ti = w.reference_time('iso')
# pr = w.pressure['press']
# vd = w.visibility_distance
# if dt == "clear sky":
# 	print('clearsky.png')
# if dt == "broken clouds":
# 	print('brokencloud.png')
# if dt == "scattered clouds":
# 	print('brokencloud.png')
# if dt == "few clouds":
# 	print('fewclouds.png')
# if dt == "shower rain":
# 	print('showerrain.png')
# if dt == "rain":
# 	print('rain.png')
# if dt == "thunderstorm":
# 	print('thunderstorm.png',)
# if dt == "snow":
# 	print('snow.png')
# if dt == "	mist":
# 	print('mist.png')
# print(dt)
# 	bot.send_message(message.chat.id, f"""
# 		***Кросс-курс***\n1֏ = `{resAR[0].text}`₽
# 		1₽ = `{resRA[0].text}`֏
# 		1$ = `{resUR[0].text}`₽/`{resUA[0].text}`֏
# 		1€ = `{resER[0].text}`₽/`{resEA[0].text}`֏
# 		***Топ 5 банков с самым выгодным курсом обмена***:
# 		Доллара:
# 		[Юнибанк/Армения](https://www.google.com/search?q=Юнибанк+Армения):
# 		🇺🇸:{result1[5].text}֏ 🇪🇺: `{result1[7].text}`֏ 🇷🇺: `{result1[9].text}`֏
# 		[Акба банк](https://www.google.com/search?q=Акба+Банк):
# 		🇺🇸:`{result2[5].text}`֏ 🇪🇺: `{result2[7].text}`֏ 🇷🇺: `{result2[9].text}`֏
# 		[Арарат Банк](https://www.google.com/search?q=Арарат+Банк):
# 		🇺🇸:`{result3[5].text}`֏ 🇪🇺: `{result3[7].text}`֏ 🇷🇺: `{result3[9].text}`֏
# 		[ВТБ Армения](https://www.google.com/search?q=ВТБ+Армения):
# 		🇺🇸:`{result4[5].text}`֏ 🇪🇺: `{result4[7].text}`֏ 🇷🇺: `{result4[9].text}`֏
# 		[Арцахбанк](https://www.google.com/search?q=Арцахбанк):
# 		🇺🇸:`{result5[5].text}`֏ 🇪🇺: `{result5[7].text}`֏ 🇷🇺: `{result5[9].text}`֏

# 		[Bitcoin](https://www.google.com/search?q=bitcoin):
# 		🇺🇸:`{resBU[0].text}`$ 🇦🇲:`{resBA[0].text}`֏ 🇷🇺:`{resBR[0].text}`₽
# 		""",parse_mode='Markdown')





















# reqpw = requests.get(f"https://api.openweathermap.org/data/2.5/weather?q=Yerevan&appid=d99de04f9c4d8f439d5149e9d7c689ff&units=metric")
# datapw = reqpw.json()
# # weatherargs = datapw["weather"]["description"]
# pprint(datapw)

# r = requests.get(
# 	f"https://api.openweathermap.org/data/2.5/weather?q=Yerevan&appid=d99de04f9c4d8f439d5149e9d7c689ff&units=metric"
# )
# data = r.json()
# print(data)

# args = data["weather"]["icon"]
# cur_weather = data["main"]["temp"]
# humidity = data["main"]["humidity"]
# pressure = data["main"]["pressure"]
# wind = data["wind"]["speed"]


# print(args)
	  
# //*[@id="forecast_list_ul"]/table/tbody/tr/td[2]/b[2]/i

# site = 'https://openweathermap.org/find?q=Ереван'
# headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.84 Safari/537.36 OPR/85.0.4341.72'}
# pars = requests.get(site, headers=headers)
# soup = bs(pars.content, 'html.parser')
# result = soup.find("div", {"class": "tab-pane active"})
# print(result)
# # soup2 = bs(pars.content, 'html.parser')
# # result2 = soup2.find("tr", {"id": "f3ffb6cf-dbb6-4d43-b49c-f6d71350d7fb"}).find_all("td")
# # soup3 = bs(pars.content, 'html.parser')
# # result3 = soup3.find("tr", {"id": "5ee70183-87fe-4799-802e-ef7f5e7323db"}).find_all("td")
# # soup4 = bs(pars.content, 'html.parser')
# # result4 = soup4.find("tr", {"id": "69460818-02ec-456e-8d09-8eeff6494bce"}).find_all("td")
# # soup5 = bs(pars.content, 'html.parser')
# # result5 = soup5.find("tr", {"id": "e1a68c2e-bc47-4f58-afd2-3b80a8465b14"}).find_all("td")
# # print(result1[5].text, result1[7].text, result1[9].text)
# # print(result2[5].text, result2[7].text, result2[9].text)
# # print(result3[5].text, result3[7].text, result3[9].text)
# # print(result4[5].text, result4[7].text, result4[9].text)
# # print(result5[5].text, result5[7].text, result5[9].text)
